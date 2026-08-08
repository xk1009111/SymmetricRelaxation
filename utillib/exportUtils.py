import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import math
import utillib.i18n as i18n
from utillib.mylib import get_distance_point_point, angle_by_three_points


'''
    椭圆拟合数据表
'''
def ellipse(excelName, cells, lineOfCell, currentTimes):

    workbook=openpyxl.Workbook()
    sheet =workbook.active

    # 添加表头
    # Add a header
    ellipse_headings = get_ellipse_headings()
    for i in range(len(ellipse_headings)):
        sheet.cell(row=1,column=i+1,value=ellipse_headings[i]).alignment=align
        if i<26:
            sheet.column_dimensions[chr(65+i)].width=18
        else:
            sheet.column_dimensions['A'+chr(65+i-26)].width=18

    i=0
    for c in cells:
        # if not c.ok:
        #     continue

        l=len(c.points)
        totalEdge=0
        perimeter=0
        # 判断是否存在相邻细胞
        # To see if there are adjacent cells
        for j in range(l):
            perimeter+=get_distance_point_point(c.points[j-1], c.points[j])
            string='{0}-{1}-{2}-{3}'.format(c.points[j][0],c.points[j][1],c.points[j-1][0],c.points[j-1][1])
            if string in lineOfCell:
                totalEdge+=len(lineOfCell[string].points)


        sheet.cell(row=2+i,column=1,value=c.no).alignment=align

        ellipse_data = c.data

        sheet.cell(row=2+i,column=2,value=str('('+str(ellipse_data['cp'].x)+','+str(ellipse_data['cp'].y)+')')).alignment=align
        sheet.cell(row=2+i,column=3,value=ellipse_data['a']).alignment=align
        sheet.cell(row=2+i,column=4,value=ellipse_data['b']).alignment=align
        sheet.cell(row=2+i,column=5,value=ellipse_data['angle']).alignment=align

        sheet.cell(row=2+i,column=6,value=totalEdge).alignment=align
        sheet.cell(row=2+i,column=7,value=l).alignment=align
        sheet.cell(row=2+i,column=8,value=perimeter).alignment=align
        sheet.cell(row=2+i,column=9,value=c.area).alignment=align

        # ================= [修改后的层数写入逻辑] =================
        # 1. 获取最初母细胞层数
        original_layer = getattr(c, 'original_layer', c.layer)

        # 2. 写入最初母细胞层数到第 10 列
        sheet.cell(row=2+i,column=10,value=original_layer).alignment=align

        # 3. 写入当前细胞层数到第 11 列
        sheet.cell(row=2+i,column=11,value=c.layer).alignment=align

        # 4. 整个循环体内【只保留这一个】i += 1，表示该细胞处理完毕，下一行准备写下一个细胞
        i += 1
        # =========================================================

    sheet.column_dimensions['B'].width = 45.0
    workbook.save('{0}_{1}.xlsx'.format(excelName, currentTimes))
    return True


def edgeangle(excelName, cells, lineOfCell, currentTimes):

    workbook=openpyxl.Workbook()
    sheet =workbook.active

    # 添加表头
    # Add a header
    edgeangle_headings = get_edgeangle_headings()
    for i in range(len(edgeangle_headings)):
        sheet.cell(row=1, column=i+1, value=edgeangle_headings[i]).alignment=align
        if i < 26:
            sheet.column_dimensions[chr(65+i)].width=18
        else:
            sheet.column_dimensions['A'+chr(65+i-26)].width=18

    i=0
    for c in cells:
        # if not c.ok:
        #     continue

        l=len(c.points)
        totalEdge=0
        # 判断是否存在相邻细胞
        # To see if there are adjacent cells
        for j in range(l):
            string='{0}-{1}-{2}-{3}'.format(c.points[j][0],c.points[j][1],c.points[j-1][0],c.points[j-1][1])
            if string in lineOfCell:
                totalEdge+=len(lineOfCell[string].points)

        for j in range(l):
            sheet.cell(row=2+i,column=1,value=c.no).alignment=align
            sheet.cell(row=2+i,column=2,value=len(c.points)).alignment=align
            sheet.cell(row=2+i,column=3,value=totalEdge).alignment=align

            p1 = c.points[j-1]
            p = c.points[j]
            p2 = c.points[(j+1)%l]

            angle = angle_by_three_points(p1, p, p2)
            d1 = get_distance_point_point(p1, p)
            d2 = get_distance_point_point(p, p2)

            sheet.cell(row=2+i,column=4,value=math.degrees(angle)).alignment=align
            sheet.cell(row=2+i,column=5,value=d1).alignment=align
            sheet.cell(row=2+i,column=6,value=d2).alignment=align
            sheet.column_dimensions[get_column_letter(4)].width = 20.0
            sheet.column_dimensions[get_column_letter(5)].width = 20.0
            sheet.column_dimensions[get_column_letter(6)].width = 20.0

            # 1. 获取最初母细胞层数
            original_layer = getattr(c, 'original_layer', c.layer)

            # 2. 写入最初母细胞层数 (对应表头第7列)
            sheet.cell(row=2+i, column=7, value=original_layer).alignment = align

            # 3. 写入当前细胞层数 (对应表头第8列)
            sheet.cell(row=2+i, column=8, value=c.layer).alignment = align

            # 4. 游标 +1，准备写入该细胞的下一个角，或者下一个细胞
            i += 1

    sheet.column_dimensions[get_column_letter(7)].width = 15.0
    workbook.save('{0}_{1}.xlsx'.format(excelName, currentTimes))
    return True


def export_ME_MA(excelName, cells, currentTimes):
    """
    导出 ME (边缘边长) 和 MA (边缘角) 到单独的 Excel 表格
    按照老师要求：将一个边缘细胞的两个边缘角(MA)和一个边缘边(ME)以及层数放在一行
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "ME_MA_Statistics"

    # 设置新的扁平化表头
    me_ma_headings = i18n.languages[i18n.current_language].get('me_ma_headings', 
        ["细胞序号", "边缘边长(ME)", "边缘角1(MA1)", "边缘角2(MA2)", "最初母细胞层数", "当前细胞层数"])
    headings = me_ma_headings
    for i, h in enumerate(headings):
        sheet.cell(row=1, column=i+1, value=h).alignment = align
        sheet.column_dimensions[get_column_letter(i+1)].width = 18

    # 1. 统计所有边的出现次数，以识别边缘边 (ME)
    edge_counts = {}
    for c in cells:
        l = len(c.points)
        for j in range(l):
            p1 = tuple(c.points[j-1])
            p2 = tuple(c.points[j])
            edge_key = tuple(sorted((p1, p2)))
            edge_counts[edge_key] = edge_counts.get(edge_key, 0) + 1

    # 提取所有边缘边 (只出现一次的边)
    boundary_edges = {k for k, v in edge_counts.items() if v == 1}

    row_idx = 2
    for c in cells:
        # 按照规定，边缘细胞为第一层，向内递增。这里只处理边缘细胞 (Layer 1)
        if c.layer != 1:
            continue

        mes = []
        mas = []

        l = len(c.points)
        for j in range(l):
            p_prev = c.points[j-1]
            p_curr = c.points[j]
            p_next = c.points[(j+1)%l]

            edge_prev_key = tuple(sorted((tuple(p_prev), tuple(p_curr))))
            edge_next_key = tuple(sorted((tuple(p_curr), tuple(p_next))))

            # 记录边缘边长
            if edge_next_key in boundary_edges:
                dist = get_distance_point_point(p_curr, p_next)
                mes.append(dist)

            # 记录边缘角 (如果顶点连接的任意一条边是边缘边)
            if edge_prev_key in boundary_edges or edge_next_key in boundary_edges:
                angle = angle_by_three_points(p_prev, p_curr, p_next)
                mas.append(math.degrees(angle))

        # 提取数据：标准的边缘细胞(不在死角)通常会有1条边缘边和2个边缘角
        me_val = mes[0] if len(mes) > 0 else "N/A"
        ma1_val = mas[0] if len(mas) > 0 else "N/A"
        ma2_val = mas[1] if len(mas) > 1 else "N/A"

        original_layer = getattr(c, 'original_layer', c.layer)

        sheet.cell(row=row_idx, column=1, value=c.no).alignment = align
        sheet.cell(row=row_idx, column=2, value=me_val).alignment = align
        sheet.cell(row=row_idx, column=3, value=ma1_val).alignment = align
        sheet.cell(row=row_idx, column=4, value=ma2_val).alignment = align
        sheet.cell(row=row_idx, column=5, value=original_layer).alignment = align
        sheet.cell(row=row_idx, column=6, value=c.layer).alignment = align
        row_idx += 1

    filename = '{0}_ME_MA_{1}.xlsx'.format(excelName, currentTimes)
    workbook.save(filename)
    return True


def create(excelName, cells, lineOfCell, currentTimes=0):
    # 依次调用三个导出函数：椭圆数据、边角数据、ME/MA数据
    res_me_ma = export_ME_MA(excelName, cells, currentTimes)

    # 调用原有的 ellipse 和 edgeangle 函数
    if ellipse("ellipse", cells, lineOfCell, currentTimes) \
        and edgeangle("edgeAngle", cells, lineOfCell, currentTimes) \
        and res_me_ma:
        return True
    else:
        return False


def get_ellipse_headings():
    return i18n.languages[i18n.current_language].get('ellipse_headings', [
        "细胞序号",
        "椭圆中心点",
        "长半轴",
        "短半轴",
        "长半轴与x轴的夹角",
        "相邻细胞边数和",
        "细胞边数",
        "细胞周长",
        "细胞面积",
        "最初母细胞层数",
        "当前细胞层数"
    ])

def get_edgeangle_headings():
    return i18n.languages[i18n.current_language].get('edgeangle_headings', [
        "细胞序号",
        "边数",
        "相邻细胞边数和",
        "内角",
        "夹边1",
        "夹边2",
        "最初母细胞层数",
        "当前细胞层数"
    ])

align=Alignment(horizontal='center',vertical='center',wrap_text=True)
